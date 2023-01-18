
import openpyxl
import webbrowser
import pyautogui
import time

path = "gif.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
#msg11 ="Respected Teacher this is the *reference articulate file* and don't copy the content as it is  and don't share with any one .its must be kept  confidential, open the following link :    "
#msg11="Good afternoon, madam/sir. *If you have finished articulating work, please send it to me via WhatsApp*. Please disregard if you are sent"
#msg12="""Good evening, sir/madam.  *The Zoom meeting was postponed to this evening*. We will send the timings and login credentials very soon. Thank you ."""
#msg123 = "Sri Chaitanya Sri Chaitanya is inviting you to a scheduled Zoom meeting.          Topic: *Infinity Learn's Articulate work*                                                                       Time: *Oct 13, 2022 04:30 PM India*                                                                         Join Zoom Meeting                                           https://srichaitanyagroup.zoom.us/j/3265584677?pwd=eEg3TWFJWDdZV0kreFZjUFpkcTEwUT09                                                                       Meeting ID: 326 558 4677                                                                      Passcode: 12345"
#msg22 = "*Dear Teachers, The Zoom meeting has been rescheduled until tomorrow. I'll let Timings know shortly...*"
msg12="*  Good afternoon ,I would like to remind you that it’s time to submit the *Articulate  work* as soon as possible because it’s already been late .I hope you will submit it early. Please disregard if already sent. Thank you"
#msg22 ="https://docs.google.com/document/d/1gZ_zuk7j7yvhVOkwpYYMOo2nROzrhtSIsw2mftQ7bMQ/edit?usp=sharing"
for i in range(2,37):
    num = sheet_obj.cell(row = i, column = 11)
    msg = sheet_obj.cell(row = i, column = 6)
    num1=str(num.value)
    msg1=str(msg.value)
    url="https://web.whatsapp.com/send?phone="+num1+"&text="+"*Please Ignore, If You have Already Sent*"
    webbrowser.open(url)
    time.sleep(10)
    pyautogui.press("enter")
    time.sleep(5)
print("task completed")